import io
import re
from copy import copy
from datetime import date
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator


APP_TITLE = "Purchase Order to Working File Dashboard"
BRAND = "@BAJRABHANU"

WORKING_SHEET_CANDIDATES = ["Working", "working"]
MASTER2_SHEET_CANDIDATES = ["master 2", "Master 2", "Master-2", "master-2"]

OUTPUT_FILENAME = "Filled_Working_File.xlsx"


st.set_page_config(page_title=APP_TITLE, page_icon="📦", layout="wide")

st.markdown(
    """
    <style>
        .main-title {
            padding: 18px 22px;
            border-radius: 18px;
            background: linear-gradient(135deg, #0f172a, #1d4ed8, #f97316);
            color: white;
            box-shadow: 0 14px 35px rgba(15, 23, 42, 0.22);
            margin-bottom: 18px;
        }
        .main-title h1 { margin: 0; font-size: 30px; }
        .main-title p { margin: 6px 0 0 0; font-size: 15px; opacity: 0.95; }
        .metric-card {
            padding: 16px;
            border-radius: 16px;
            border: 1px solid #e5e7eb;
            background: #ffffff;
            box-shadow: 0 8px 24px rgba(15, 23, 42, 0.08);
        }
        .metric-label { color:#64748b; font-size:13px; font-weight:700; }
        .metric-value { color:#0f172a; font-size:28px; font-weight:800; margin-top:4px; }
        .note-box {
            padding: 14px 16px;
            border-radius: 14px;
            background: #fff7ed;
            border: 1px solid #fed7aa;
            color:#7c2d12;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <div class='main-title'>
        <h1>📦 Purchase Order Automation Dashboard</h1>
        <p>
        Upload your formulated working file + multiple Flipkart PO files, then download the filled Excel file.
        {BRAND}
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def clean_header(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean_text(value).upper())


def to_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)) and not pd.isna(value):
        return value

    text = clean_text(value)
    if not text:
        return None

    text = text.replace(",", "")
    text = re.sub(r"\bINR\b|%", "", text, flags=re.IGNORECASE).strip()

    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None

    number = float(match.group(0))
    return int(number) if number.is_integer() else number


def read_po_as_dataframe(uploaded_file) -> pd.DataFrame:
    file_bytes = uploaded_file.getvalue()
    name = uploaded_file.name.lower()

    if name.endswith(".xls"):
        engine = "xlrd"
    else:
        engine = "openpyxl"

    return pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=0,
        header=None,
        dtype=object,
        engine=engine,
    )


def find_label_value(df: pd.DataFrame, label: str) -> str:
    """
    Finds label in any cell and returns the next non-empty cell on same row.
    Example:
    PO#                 FABWL08088224
    SHIPPED TO ADDRESS  Flipkart India Pvt Ltd...
    """
    target = clean_header(label)

    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            if clean_header(df.iat[r, c]) == target:
                for cc in range(c + 1, df.shape[1]):
                    value = clean_text(df.iat[r, cc])
                    if value:
                        return value

    return ""


def find_order_header_row(df: pd.DataFrame) -> Optional[int]:
    for r in range(df.shape[0]):
        headers = [clean_header(x) for x in df.iloc[r].tolist()]

        if "FSNISBN13" in headers and "QUANTITY" in headers:
            return r

    return None


def parse_po_file(uploaded_file) -> Tuple[pd.DataFrame, List[str]]:
    errors: List[str] = []

    try:
        df = read_po_as_dataframe(uploaded_file)
    except Exception as exc:
        return pd.DataFrame(), [f"{uploaded_file.name}: unable to read file. {exc}"]

    po_no = find_label_value(df, "PO#")
    shipped_to_address = find_label_value(df, "SHIPPED TO ADDRESS")

    if not po_no:
        all_text = " ".join(clean_text(x) for x in df.to_numpy().flatten())

        match = re.search(
            r"PURCHASE\s+ORDER\s*#\s*([A-Z0-9-]+)",
            all_text,
            flags=re.IGNORECASE,
        )

        if match:
            po_no = match.group(1).strip()

    header_row = find_order_header_row(df)

    if header_row is None:
        return pd.DataFrame(), [f"{uploaded_file.name}: order details table not found."]

    header_map: Dict[str, int] = {}

    for c, value in enumerate(df.iloc[header_row].tolist()):
        header_map[clean_header(value)] = c

    fsn_col = header_map.get("FSNISBN13")
    qty_col = header_map.get("QUANTITY")
    uom_col = header_map.get("UOM")
    title_col = header_map.get("TITLE")
    price_col = header_map.get("SUPPLIERPRICE")
    taxable_col = header_map.get("TAXABLEVALUE")
    req_date_col = header_map.get("REQUIREDBYDATE")

    if fsn_col is None or qty_col is None:
        return pd.DataFrame(), [f"{uploaded_file.name}: FSN/ISBN13 or Quantity column missing."]

    rows: List[Dict[str, object]] = []

    for r in range(header_row + 1, df.shape[0]):
        fsn = clean_text(df.iat[r, fsn_col])
        qty = to_number(df.iat[r, qty_col])

        if not fsn:
            continue

        if "TOTAL" in fsn.upper() or "IMPORTANT" in fsn.upper():
            break

        if qty is None:
            continue

        rows.append(
            {
                "PO#": po_no,
                "SHIPPED TO ADDRESS": shipped_to_address,
                "FSN/ISBN13": fsn,
                "Quantity": qty,
                "UOM": clean_text(df.iat[r, uom_col]) if uom_col is not None else "",
                "Title": clean_text(df.iat[r, title_col]) if title_col is not None else "",
                "Supplier Price": to_number(df.iat[r, price_col]) if price_col is not None else None,
                "Taxable Value": to_number(df.iat[r, taxable_col]) if taxable_col is not None else None,
                "Required by Date": clean_text(df.iat[r, req_date_col]) if req_date_col is not None else "",
                "Source File": uploaded_file.name,
            }
        )

    if not rows:
        errors.append(f"{uploaded_file.name}: no item rows found.")

    if not po_no:
        errors.append(f"{uploaded_file.name}: PO number not found; PO# will remain blank.")

    if not shipped_to_address:
        errors.append(f"{uploaded_file.name}: shipped to address not found; Master-2 address will remain blank.")

    return pd.DataFrame(rows), errors


def find_sheet_name(wb, candidates: List[str]) -> str:
    existing_map = {name.strip().lower(): name for name in wb.sheetnames}

    for candidate in candidates:
        key = candidate.strip().lower()
        if key in existing_map:
            return existing_map[key]

    raise ValueError(f"Required sheet not found. Expected one of: {', '.join(candidates)}")


def copy_cell_style_and_formula(ws, source_row: int, target_row: int, col: int) -> None:
    src = ws.cell(source_row, col)
    dst = ws.cell(target_row, col)

    if src.has_style:
        dst._style = copy(src._style)

    if src.number_format:
        dst.number_format = src.number_format

    if src.alignment:
        dst.alignment = copy(src.alignment)

    if src.border:
        dst.border = copy(src.border)

    if src.fill:
        dst.fill = copy(src.fill)

    if src.font:
        dst.font = copy(src.font)

    if src.protection:
        dst.protection = copy(src.protection)

    if isinstance(src.value, str) and src.value.startswith("="):
        try:
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        except Exception:
            dst.value = src.value
    else:
        dst.value = src.value


def copy_row_format_and_formulas(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        copy_cell_style_and_formula(ws, source_row, target_row, col)

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def fill_working_sheet(ws, extracted_df: pd.DataFrame, po_date_value: Optional[date]) -> None:
    max_col = ws.max_column
    required_rows = max(2, len(extracted_df) + 1)

    for target_row in range(2, required_rows + 1):
        source_row = 2
        copy_row_format_and_formulas(ws, source_row, target_row, max_col)

    for r in range(2, max(ws.max_row, required_rows) + 1):
        ws.cell(r, 1).value = None
        ws.cell(r, 2).value = None
        ws.cell(r, 3).value = None

    for idx, row in extracted_df.reset_index(drop=True).iterrows():
        excel_row = idx + 2

        ws.cell(excel_row, 1).value = row.get("PO#", "")
        ws.cell(excel_row, 2).value = row.get("FSN/ISBN13", "")
        ws.cell(excel_row, 3).value = row.get("Quantity", "")

        if po_date_value:
            ws.cell(excel_row, 17).value = po_date_value.strftime("%d.%m.%Y")


def fill_master2_sheet(ws, po_master_df: pd.DataFrame) -> None:
    """
    Fill Master-2 orange/manual area.

    Expected Master-2 structure:
    A = Origin Warehouse formula
    B = PO#
    C = SHIPPED TO ADDRESS
    D = Sap Cde formula
    E = Plant Code formula
    F = Origin Warehouse formula/value

    Only B and C are overwritten.
    A, D, E, F formulas are preserved/copied.
    Product master columns R:U are untouched.
    """

    required_rows = max(2, len(po_master_df) + 1)

    # Copy formulas/styles only for A:F.
    for target_row in range(2, required_rows + 1):
        source_row = 2

        for col in range(1, 7):
            copy_cell_style_and_formula(ws, source_row, target_row, col)

        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    # Clear only manual input columns B:C.
    # Do not touch R:U product master.
    for r in range(2, max(ws.max_row, required_rows) + 1):
        ws.cell(r, 2).value = None
        ws.cell(r, 3).value = None

    for idx, row in po_master_df.reset_index(drop=True).iterrows():
        excel_row = idx + 2

        ws.cell(excel_row, 2).value = row.get("PO#", "")
        ws.cell(excel_row, 3).value = row.get("SHIPPED TO ADDRESS", "")


def add_import_summary_sheet(wb, extracted_df: pd.DataFrame, po_master_df: pd.DataFrame) -> None:
    if "PO Import Summary" in wb.sheetnames:
        del wb["PO Import Summary"]

    summary = wb.create_sheet("PO Import Summary")

    summary.append(["PO Import Summary"])
    summary.append([])
    summary.append(["Total PO Lines", len(extracted_df)])
    summary.append(["Unique PO Count", po_master_df["PO#"].nunique()])
    summary.append(["Total Quantity", extracted_df["Quantity"].sum()])
    summary.append([])

    summary.append(["PO-wise Master-2 Data"])
    summary.append(["PO#", "SHIPPED TO ADDRESS"])

    for _, row in po_master_df.iterrows():
        summary.append(
            [
                row.get("PO#", ""),
                row.get("SHIPPED TO ADDRESS", ""),
            ]
        )

    summary.append([])
    summary.append(["Line-wise Working Sheet Data"])

    detail_headers = list(extracted_df.columns)
    summary.append(detail_headers)

    for _, row in extracted_df.iterrows():
        summary.append([row.get(col, "") for col in detail_headers])

    for col_cells in summary.columns:
        max_len = max(len(clean_text(cell.value)) for cell in col_cells)
        summary.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 45)

    summary.freeze_panes = "A8"


def fill_working_file(template_file, extracted_df: pd.DataFrame, po_date_value: Optional[date]) -> bytes:
    wb = load_workbook(template_file)

    working_sheet_name = find_sheet_name(wb, WORKING_SHEET_CANDIDATES)
    master2_sheet_name = find_sheet_name(wb, MASTER2_SHEET_CANDIDATES)

    ws_working = wb[working_sheet_name]
    ws_master2 = wb[master2_sheet_name]

    po_master_df = (
        extracted_df[["PO#", "SHIPPED TO ADDRESS"]]
        .drop_duplicates(subset=["PO#"], keep="first")
        .reset_index(drop=True)
    )

    fill_working_sheet(
        ws=ws_working,
        extracted_df=extracted_df,
        po_date_value=po_date_value,
    )

    fill_master2_sheet(
        ws=ws_master2,
        po_master_df=po_master_df,
    )

    add_import_summary_sheet(
        wb=wb,
        extracted_df=extracted_df,
        po_master_df=po_master_df,
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


with st.sidebar:
    st.header("Upload Files")

    working_file = st.file_uploader(
        "Upload formulated working file (.xlsx)",
        type=["xlsx"],
    )

    po_files = st.file_uploader(
        "Upload Purchase Order files (.xls / .xlsx)",
        type=["xls", "xlsx"],
        accept_multiple_files=True,
    )

    st.divider()

    use_po_date = st.checkbox(
        "Overwrite PO Date in Working sheet column Q",
        value=False,
    )

    po_date_value = st.date_input(
        "PO Date",
        value=date.today(),
        disabled=not use_po_date,
    )

    process_clicked = st.button(
        "🚀 Process PO Files",
        type="primary",
        use_container_width=True,
    )


st.markdown(
    """
    <div class='note-box'>
    This app fills:
    <br><br>
    <b>Working sheet</b>: PO#, FSN/ISBN13, Quantity
    <br>
    <b>Master-2 / master 2 sheet</b>: PO#, SHIPPED TO ADDRESS
    <br><br>
    Existing formulas, master data and product mapping columns are preserved.
    </div>
    """,
    unsafe_allow_html=True,
)


if process_clicked:
    if working_file is None:
        st.error("Please upload the formulated working file first.")
        st.stop()

    if not po_files:
        st.error("Please upload at least one purchase order file.")
        st.stop()

    all_frames: List[pd.DataFrame] = []
    all_errors: List[str] = []

    with st.spinner("Reading PO files and preparing output workbook..."):
        for po in po_files:
            po_df, errors = parse_po_file(po)

            if not po_df.empty:
                all_frames.append(po_df)

            all_errors.extend(errors)

        if not all_frames:
            st.error("No valid PO item rows could be extracted.")

            for err in all_errors:
                st.warning(err)

            st.stop()

        extracted = pd.concat(all_frames, ignore_index=True)

        extracted = extracted[
            [
                "PO#",
                "SHIPPED TO ADDRESS",
                "FSN/ISBN13",
                "Quantity",
                "UOM",
                "Title",
                "Supplier Price",
                "Taxable Value",
                "Required by Date",
                "Source File",
            ]
        ]

        output_bytes = fill_working_file(
            template_file=working_file,
            extracted_df=extracted,
            po_date_value=po_date_value if use_po_date else None,
        )

    po_master_preview = (
        extracted[["PO#", "SHIPPED TO ADDRESS"]]
        .drop_duplicates(subset=["PO#"], keep="first")
        .reset_index(drop=True)
    )

    c1, c2, c3, c4 = st.columns(4)

    with c1:
        st.markdown(
            f"""
            <div class='metric-card'>
                <div class='metric-label'>PO Files</div>
                <div class='metric-value'>{len(po_files)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with c2:
        st.markdown(
            f"""
            <div class='metric-card'>
                <div class='metric-label'>Rows Extracted</div>
                <div class='metric-value'>{len(extracted)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with c3:
        st.markdown(
            f"""
            <div class='metric-card'>
                <div class='metric-label'>Total Quantity</div>
                <div class='metric-value'>{extracted['Quantity'].sum():,.0f}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with c4:
        st.markdown(
            f"""
            <div class='metric-card'>
                <div class='metric-label'>Unique PO#</div>
                <div class='metric-value'>{extracted['PO#'].nunique()}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.subheader("Preview: Working Sheet Data")
    st.dataframe(
        extracted[
            [
                "PO#",
                "FSN/ISBN13",
                "Quantity",
                "UOM",
                "Title",
                "Supplier Price",
                "Taxable Value",
                "Required by Date",
                "Source File",
            ]
        ],
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Preview: Master-2 Data")
    st.dataframe(
        po_master_preview,
        use_container_width=True,
        hide_index=True,
    )

    if all_errors:
        with st.expander("Warnings / files needing review"):
            for err in all_errors:
                st.warning(err)

    st.download_button(
        label="⬇️ Download Filled Working Excel File",
        data=output_bytes,
        file_name=OUTPUT_FILENAME,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

else:
    st.info("Upload the working file and PO files from the left side, then click Process PO Files.")
